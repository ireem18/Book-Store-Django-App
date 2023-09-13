from django.contrib import messages
from django.http import JsonResponse, HttpResponseRedirect
from django.db.models import CharField, Value, Count
from django.db.models.functions import Concat
from django.shortcuts import render, redirect, get_object_or_404

from django.db.models import Q
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

from django.contrib.auth.decorators import login_required, permission_required
from django.template.loader import render_to_string

from writer.models import Writer
from .forms import BookForm, SearchForm
from .models import Book

@login_required(login_url="login")
@permission_required('books.can_view_book_list', raise_exception=True)
def chart_board(request):
    writer_info = Book.objects.active().values('writer').annotate(full_name=Concat("writer__name", Value(" "), "writer__surname",
                                                                output_field=CharField()), total=Count('writer')).values('full_name', 'total')
    publisher_info = Writer.objects.active().values('publisher').annotate(total=Count('publisher')).values('publisher__name', 'total')
    book_info = Book.objects.active().values('name', 'count')
    context = {'writerInfo': writer_info, 'publisherInfo': publisher_info, 'bookInfo': book_info}
    return render(request, "chart_board.html", context)



@login_required(login_url="login")
@permission_required('books.can_view_book_list', raise_exception=True)
def book_list(request):
    try:
        books = Book.objects.active()
        form = SearchForm(request.POST)

        if request.method == 'POST':
            if form.is_valid():
                query = form.cleaned_data['query']
                books = books.filter(Q(publisher__name__icontains=query) | Q(writer__name__icontains=query) |\
                                     Q(writer__surname__icontains=query) | Q(name__icontains=query))

        paginator = Paginator(books, 3)
        page = request.GET.get('page', 1)
        try:
            books = paginator.page(page)
        except PageNotAnInteger:
            books = paginator.page(1)
        except EmptyPage:
            books = paginator.page(paginator.num_pages)

        content = {
            'books': books,
            'form': form
        }
        return render(request, "book_list.html", content)
    except Exception as e:
        print('Book List Error!', str(e))
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))


@login_required(login_url="login")
def save_book_form(request, form, template_name):
    try:
        data = dict()
        if request.method == 'POST':
            if form.is_valid():
                form.save()
                data['form_is_valid'] = True
                books = Book.objects.active()
                data['html_book_list'] = render_to_string('include_book_list.html', {
                    'books': books
                })
            else:
                data['form_is_valid'] = False
        context = {'form': form, 'formPage': True}
        data['html_form'] = render_to_string(template_name, context, request=request)
        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({"success": False})


@login_required(login_url="login")
@permission_required('books.can_add_book', raise_exception=True)
def add_book(request):
    if request.method == 'POST':
        form = BookForm(request.POST)
    else:
        form = BookForm()
    return save_book_form(request, form, 'book_create.html')


@login_required(login_url="login")
@permission_required('books.can_edit_book', raise_exception=True)
def edit_book(request, id):
    book = get_object_or_404(Book, pk=id)
    if request.method == 'POST':
        form = BookForm(request.POST, instance=book)
    else:
        form = BookForm(instance=book)
    return save_book_form(request, form, 'book_update.html')


@login_required(login_url="login")
@permission_required('books.can_delete_book', raise_exception=True)
def delete_book(request, id):
    try:
        book = Book.objects.get(id=id)
        if request.method == 'POST':
            book.deactive()
            messages.success(request, 'Book Deleted')
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
        else:
            context = {'book': book, 'formPage': True}
            html_form = render_to_string('book_delete.html', context, request=request)
            return JsonResponse({'html_form': html_form})
    except Book.DoesNotExist:
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))


@login_required(login_url="login")
@permission_required('books.can_view_book_list', raise_exception=True)
def export_book_list(request):
    import openpyxl
    import datetime
    from django.http import HttpResponse

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="book_list.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Book List'

    # Write header row
    header = ['Publisher Name', 'Writer Name', 'Name', 'Subject', 'ISBN', 'Count', 'Page Count', 'Publisher Date']
    for col_num, column_title in enumerate(header, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title
        cell.font = openpyxl.styles.Font(color='FF0000', bold=True)
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="DDDDDD")

    # Write data rows
    queryset = Book.objects.active().annotate(full_name=Concat("writer__name", Value(" "), "writer__surname", output_field=CharField()))
    queryset = queryset.values_list('publisher__name', 'full_name', 'name', 'subject', 'isbn', 'count', 'page_count',
                                    'publisher_date')

    for row_num, row in enumerate(queryset, 1):
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num + 1, column=col_num)
            if isinstance(cell_value, datetime.datetime):
                cell_value = cell_value.strftime('%Y-%m-%d')
            cell.value = cell_value
            cell.font = openpyxl.styles.Font(bold=True)

    workbook.save(response)
    return response